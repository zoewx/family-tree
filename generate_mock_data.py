import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Family Tree"

# Header row
headers = ["row_id", "firstName", "lastName", "gender", "birthDate", "deathDate", "bio", "father_row_id", "mother_row_id", "spouse_row_id", "generation", "englishName"]
ws.append(headers)

# 三代家庭数据 (Three generations - 王家)
data = [
    # Generation 0 - 爷爷奶奶
    [1, "大明", "王", "男", "1940-03-15", "2015-06-20", "家族长辈，曾任教师",       None, None, 2,    0, "David Wang"],
    [2, "秀英", "李", "女", "1942-08-22", None,         "家族祖母，擅长烹饪",       None, None, 1,    0, "Susan Li"],

    # Generation 0 - 外公外婆
    [3, "建国", "张", "男", "1938-11-03", "2018-02-10", "退休工程师",              None, None, 4,    0, "James Zhang"],
    [4, "玉兰", "刘", "女", "1941-05-17", None,         "退休医生",                None, None, 3,    0, "Grace Liu"],

    # Generation 1 - 爸爸妈妈
    [5, "志强", "王", "男", "1965-07-10", None,         "软件工程师，家中长子",      1,    2,    6,    1, "Kevin Wang"],
    [6, "美玲", "张", "女", "1968-01-25", None,         "中学教师",                3,    4,    5,    1, "Meilin Zhang"],

    # Generation 1 - 叔叔婶婶
    [7, "志勇", "王", "男", "1970-04-18", None,         "医生，家中次子",            1,    2,    8,    1, "Victor Wang"],
    [8, "晓红", "陈", "女", "1972-09-30", None,         "护士",                    None, None, 7,    1, "Helen Chen"],

    # Generation 2 - 孩子们
    [9,  "浩然", "王", "男", "1990-12-05", None,         "程序员，长子",              5,    6,    None, 2, "Ryan Wang"],
    [10, "思琪", "王", "女", "1993-06-15", None,         "设计师，长女",              5,    6,    None, 2, "Sophie Wang"],
    [11, "俊杰", "王", "男", "1995-03-20", None,         "大学生",                  7,    8,    None, 2, "Jason Wang"],
    [12, "雨萱", "王", "女", "1998-11-08", None,         "高中生",                  7,    8,    None, 2, "Emily Wang"],
]

for row in data:
    ws.append(row)

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col[0].column_letter].width = max_length + 4

# Style header
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill

wb.save("mock_family_data.xlsx")
print("Created mock_family_data.xlsx — 12 persons, 3 generations")
